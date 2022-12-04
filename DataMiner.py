from github import Github
import json
import os
import pandas as pd
import openpyxl
import csv

class ContentCallback:
    def __init__(self):
        self.contents = ''

    def content_callback(self, buf):
        self.contents = self.contents + str(buf)

filename = "nvdcve-1.1-"
filetype = ".json"
data_files = []
ACCESS_TOKEN = ""
g = Github(ACCESS_TOKEN)

r = 2

#check to see if data.xlsx exists
try:
    df = pd.read_excel("data.xlsx")
    #delete the file
    os.remove("data.xlsx")
    writer = pd.ExcelWriter('data.xlsx', engine='xlsxwriter')
    writer.close()
except:
    #if it doesn't exist, create it
    writer = pd.ExcelWriter('data.xlsx', engine='xlsxwriter')
    writer.close()

def search_github(keywords):
    list = []
    query = keywords + '+in:readme+in:description'
    print(query)
    result = g.search_repositories(query, 'stars', 'desc')
    print(f'Found {result.totalCount} repo(s)')
    for repo in result:
        list.append(repo.clone_url)
    return list

def exploitdb_searching(name):

    #parse the files_exploits.csv file
    #df = pd.read_csv("files_exploits.csv")
    try:
        with open('files_exploits.csv', 'rt') as f:
            reader = csv.reader(f, delimiter=',')
            for s in reader:
                if s[11][:13] == name:
                    row = s
        #get the description from the row
        description = row[2]
        #get the date published from the row
        date = row[3]
        #get the file from the row
        file = row[1]
        #return an array of the description, date, and file
    #except
    except Exception as e:
        description = []
        date = []
        file = []
    return [description, date, file]
    


for i in range(2002, 2022):
    data_files.append("data/" + filename + str(i) + filetype)

toWrite = (("CVE ID", "CVE Description", "CVE Published Date", "CVE Last Modified Date", "CVE CVSS Score", "CVE CVSS Severity", "CVE CVSS Vector", "CVE CWE ID", "ExploitDB URL", "ExploitDB Title", "ExploitDB Publish Date", "Github URL"))
#write headers to excel
wb = openpyxl.load_workbook(filename='data.xlsx')
sheet = wb.active
sheet.append(toWrite)
wb.save('data.xlsx')

#read through the data and add it to excel
for file in data_files:
    #print the file name
    print(file)
    #clear all used in the loop
    cve_id = ""
    cve_description = ""
    cve_published_date = ""
    cve_last_modified_date = ""
    cve_cvss_score = ""
    cve_cvss_severity = ""
    cve_cvss_vector = ""
    cve_cwe_id = ""
    exploitdb_search = None
    toWrite = ()
    sheet = None
    wb = None
    with open(file) as f:
        data = json.load(f)
        for i in range(len(data["CVE_Items"])):
            if "REJECT" in data["CVE_Items"][i]["cve"]["description"]["description_data"][0]["value"]:
                #skip this CVE
                continue
            cve_id = data["CVE_Items"][i]["cve"]["CVE_data_meta"]["ID"]
            cve_description = data["CVE_Items"][i]["cve"]["description"]["description_data"][0]["value"]
            cve_published_date = data["CVE_Items"][i]["publishedDate"]
            cve_last_modified_date = data["CVE_Items"][i]["lastModifiedDate"]
            try:
                cve_cvss_score = data["CVE_Items"][i]["impact"]["baseMetricV3"]["cvssV3"]["baseScore"]
            except:
                cve_cvss_score = data["CVE_Items"][i]["impact"]["baseMetricV2"]["cvssV2"]["baseScore"]
            try:
                cve_cvss_severity = data["CVE_Items"][i]["impact"]["baseMetricV3"]["cvssV3"]["baseSeverity"]
            except:
                cve_cvss_severity = data["CVE_Items"][i]["impact"]["baseMetricV2"]["severity"]
            try:
                cve_cvss_vector = data["CVE_Items"][i]["impact"]["baseMetricV3"]["cvssV3"]["vectorString"]
            except:
                cve_cvss_vector = data["CVE_Items"][i]["impact"]["baseMetricV2"]["cvssV2"]["vectorString"]
            cve_cwe_id = data["CVE_Items"][i]["cve"]["problemtype"]["problemtype_data"][0]["description"][0]["value"]
            cve_cwe_id = cve_cwe_id.replace("CWE-", "")
            cve_cwe_id = cve_cwe_id.replace("CVE-", "")
            cve_cwe_id = cve_cwe_id.replace(" ", "")
            cve_cwe_id = cve_cwe_id.replace("-", "")
            cve_cwe_id = cve_cwe_id.replace(":", "")
            cve_cwe_id = cve_cwe_id.replace(";", "")
            cve_cwe_id = cve_cwe_id.replace(",", "")
            cve_cwe_id = cve_cwe_id.replace(".", "")
            cve_cwe_id = cve_cwe_id.replace("(", "")
            cve_cwe_id = cve_cwe_id.replace(")", "")
            cve_cwe_id = cve_cwe_id.replace("]", "")
            cve_cwe_id = cve_cwe_id.replace("[", "")
            cve_cwe_id = cve_cwe_id.replace("{", "")
            cve_cwe_id = cve_cwe_id.replace("}", "")
            cve_cwe_id = cve_cwe_id.replace("=", "")
            cve_cwe_id = cve_cwe_id.replace("!", "")
            cve_cwe_id = cve_cwe_id.replace("@", "")
            cve_cwe_id = cve_cwe_id.replace("#", "")
            cve_cwe_id = cve_cwe_id.replace("$", "")
            cve_cwe_id = cve_cwe_id.replace("%", "")
            cve_cwe_id = cve_cwe_id.replace("^", "")
            cve_cwe_id = cve_cwe_id.replace("&", "")
            #search exploitdb for exploits
            exploitdb_search = exploitdb_searching(cve_id)
            #search github for Poc's
            ####Commented out due to GitHub API rate limiting
            #github_search = search_github(cve_id + " Poc")
            github_search = None
            
            #check results from searches
            #add to excel sheet for the year, enumerate through the list of results
            if exploitdb_search is not None and len(exploitdb_search[0]) > 0 and github_search is not None and github_search != []:
                for i in range(len(exploitdb_search)):
                    toWrite = ((cve_id, cve_description, cve_published_date, cve_last_modified_date, cve_cvss_score, cve_cvss_severity, cve_cvss_vector, cve_cwe_id, exploitdb_search[2], exploitdb_search[0], exploitdb_search[1], github_search))
                    wb = openpyxl.load_workbook(filename='data.xlsx')
                    sheet = wb.active
                    for col, item in enumerate(toWrite, sheet.min_column):
                        sheet.cell(column=col, row=r, value=item)
                    r += 1
                    wb.save('data.xlsx')
            elif exploitdb_search is not None and len(exploitdb_search[0]) > 0 and github_search is None:
                for i in range(len(exploitdb_search)):
                    toWrite = ((cve_id, cve_description, cve_published_date, cve_last_modified_date, cve_cvss_score, cve_cvss_severity, cve_cvss_vector, cve_cwe_id, exploitdb_search[2], exploitdb_search[0], exploitdb_search[1], github_search))
                    wb = openpyxl.load_workbook(filename='data.xlsx')
                    sheet = wb.active
                    for col, item in enumerate(toWrite, sheet.min_column):
                        sheet.cell(column=col, row=r, value=item)
                    r += 1
                    wb.save('data.xlsx')
            #elif github_search is not None and github_search != []:
            #    toWrite = ((cve_id, cve_description, cve_published_date, cve_last_modified_date, cve_cvss_score, cve_cvss_severity, cve_cvss_vector, cve_cwe_id, "None", "None", "None", github_search))
            #    wb = openpyxl.load_workbook(filename='data.xlsx')
            #    sheet = wb.active
            #    sheet.append(toWrite)
            #    wb.save('data.xlsx')
            else:
                toWrite = ((cve_id, cve_description, cve_published_date, cve_last_modified_date, cve_cvss_score, cve_cvss_severity, cve_cvss_vector, cve_cwe_id, "None", "None", "None", "None"))
                wb = openpyxl.load_workbook(filename='data.xlsx')
                sheet = wb.active
                for col, item in enumerate(toWrite, sheet.min_column):
                        sheet.cell(column=col, row=r, value=item)
                r += 1
                wb.save('data.xlsx')
            if r % 100 == 0:
                print("Rows written: " + str(r))
